# coding:utf-8
'''
# 希望对大家有帮助哈，请多提问题
create by yaoyz
date: 2017/01/24

openpyxl==2.6.4
'''

import xlrd
import xlwt

# workbook相关
from openpyxl.workbook import Workbook

# ExcelWriter，封装了很强大的excel写的功能
from openpyxl.writer.excel import ExcelWriter

# 一个eggache的数字转为列字母的方法
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook


class HandleExcel():
    '''Excel相关操作类'''
    def __init__(self):
        self.head_row_labels = [
            u'学生ID', u'学生姓名', u'联系方式', u'知识点ID', u'知识点名称', 'memo'
        ]

    """
        function：
            读出txt文件中的每一条记录，把它保存在list中

        Param:
            filename:  要读出的文件名

        Return:
            res_list：返回的记录的list
    """

    def read_from_file(self, filename):
        res_list = []
        file_obj = open(filename, "r")
        for line in file_obj.readlines():
            res_list.append(line.split(','))

        file_obj.close()

        return res_list

    """
        function：
            读出*.xlsx中的每一条记录，把它保存在data_dic中返回

        Param:
            excel_name:  要读出的文件名

        Return:
            data_dic：返回的记录的dict
    """

    def read_excel_with_openpyxl(self, excel_name="testexcel2007.xlsx"):
        # 读取excel2007文件
        wb = load_workbook(filename=excel_name)

        # 显示有多少张表
        print("Worksheet range(s):", wb.get_named_ranges())
        print("Worksheet name(s):", wb.get_sheet_names())

        # 取第一张表
        sheetnames = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheetnames[0])

        # 显示表名，表行数，表列数
        print("Work Sheet Titile:", ws.title)
        print("Work Sheet Rows:", ws.get_highest_row())

        print("Work Sheet Cols:", ws.get_highest_column())

        # 获取读入的excel表格的有多少行，有多少列
        row_num = ws.get_highest_row()
        col_num = ws.get_highest_column()

        print("row_num: ", row_num, " col_num: ", col_num)

        # 建立存储数据的字典
        data_dic = {}
        sign = 1

        # 把数据存到字典中
        for row in ws.rows:
            temp_list = []
            #print(  "row",row)

            for cell in row:
                print(cell.value, )
                temp_list.append(cell.value)
                print("")
            data_dic[sign] = temp_list
            sign += 1

        print(data_dic)
        return data_dic

    """
        function：
            读出*.xlsx中的每一条记录，把它保存在data_dic中返回

        Param:
            records: 要保存的，一个包含每一条记录的list
            save_excel_name:  保存为的文件名
            head_row_stu_arrive_star:

        Return:
            data_dic：返回的记录的dict

    """

    def write_to_excel_with_openpyxl(self,
                                     records,
                                     head_row,
                                     save_excel_name="save.xlsx"):

        # 新建一个workbook
        wb = Workbook()

        # 新建一个excelWriter
        #ew = ExcelWriter(workbook=wb)

        # 设置文件输出路径与名称
        dest_filename = save_excel_name.decode('utf-8')

        # 第一个sheet是ws
        ws = wb.worksheets[0]

        # 设置ws的名称
        ws.title = "range names"

        # 写第一行，标题行
        ## for h_x in range(1, len(head_row) + 1):
        ## h_col = get_column_letter(h_x)
        ## ws['%s%s' % (h_col, 1)].value = '%s' % head_row[h_x - 1]
        ## ws['A2'].value = 'xx'
        ws.append(head_row)

        # 写第二行及其以后的那些行
        for record in records:
            ws.append(record)

        # save文件
        wb.save(filename=dest_filename)

    """
        function：
            测试输出Excel内容
            读出Excel文件

        Param:
            excel_name:  要读出的Excel文件名

        Return:
          无
    """

    def read_excel(self, excel_name):
        workbook = xlrd.open_workbook(excel_name)
        print(workbook.sheet_names())

        # 获取所有sheet
        print(workbook.sheet_names())  # [u'sheet1', u'sheet2']
        sheet2_name = workbook.sheet_names()[1]

        # 根据sheet索引或者名称获取sheet内容
        sheet2 = workbook.sheet_by_index(1)  # sheet索引从0开始
        sheet2 = workbook.sheet_by_name('Sheet1')

        # sheet的名称，行数，列数
        print(sheet2.name, sheet2.nrows, sheet2.ncols)

        # 获取整行和整列的值（数组）
        rows = sheet2.row_values(3)  # 获取第四行内容
        cols = sheet2.col_values(2)  # 获取第三列内容

        print(rows)

        print(cols)

        # 获取单元格内容
        print(sheet2.cell(1, 0).value)
        print(sheet2.cell_value(1, 0))
        print(sheet2.row(1)[0].value)

        # 获取单元格内容的数据类型
        print(sheet2.cell(1, 0).ctype)

        # 通过名称获取
        return workbook.sheet_by_name(u'Sheet1')

    """
        function：
            设置单元格样式

        Param:
            name:  字体名字
            height:  字体高度
            bold:  是否大写

        Return:
            style: 返回设置好的格式对象
    """

    def set_style(self, name, height, bold=False):

        style = xlwt.XFStyle()  # 初始化样式
        font = xlwt.Font()  # 为样式创建字体
        font.name = name  #'Times New Roman'
        font.bold = bold
        font.color_index = 4
        font.height = height
        borders = xlwt.Borders()
        borders.left = 6
        borders.right = 6
        borders.top = 6
        borders.bottom = 6
        style.font = font
        style.borders = borders
        return style

    """
        function：
            按照 设置单元格样式  把计算结果由txt转变为Excel存储

        Param:
            dataset：要保存的结果数据，list存储

        Return:
            将结果保存为 excel对象中
    """

    def write_to_excel(self, dataset, save_excel_name, head_row):

        f = xlwt.Workbook()  # 创建工作簿
        # 创建第一个sheet:
        # sheet1
        count = 1
        sheet1 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
        # 首行标题：
        for p in range(len(head_row)):
            sheet1.write(0, p, head_row[p],
                         self.set_style('Times New Roman', 250, True))
            default = self.set_style(
                'Times New Roman', 200,
                False)  # define style out the loop will work

        for line in dataset:
            row_list = str(line).strip("\n").split("\t")

            for pp in range(len(str(line).strip("\n").split("\t"))):
                sheet1.write(count, pp, row_list[pp].decode('utf-8'), default)

        count += 1
        f.save(save_excel_name)  # 保存文件

    def run_main_save_to_excel_with_openpyxl(self):
        print("测试读写2007及以后的excel文件xlsx，以方便写入文件更多数据")
        print("1. 把txt文件读入到内存中，以list对象存储")
        dataset_list = self.read_from_file("test_excel.txt")
        '''test use openpyxl to handle EXCEL 2007'''
        print("2. 把文件写入到Excel表格中")
        head_row_label = self.head_row_labels
        save_name = "test_openpyxl.xlsx"
        self.write_to_excel_with_openpyxl(dataset_list, head_row_label,
                                          save_name)
        print("3.  执行完毕，由txt格式文件保存为Excel文件的任务")

    def run_main_save_to_excel_with_xlwt(self):
        print("4. 把txt文件读入到内存中，以list对象存储")
        dataset_list = self.read_from_file("test_excel.txt")
        '''test use xlwt to handle EXCEL 97-2003'''
        print("5. 把文件写入到Excel表格中")
        head_row_label = self.head_row_labels
        save_name = "test_xlwt.xls"
        self.write_to_excel_with_openpyxl(dataset_list, head_row_label,
                                          save_name)
        print("6.  执行完毕，由txt格式文件保存为Excel文件的任务")


if __name__ == '__main__':

    print("create handle Excel Object")

    obj_handle_excel = HandleExcel()

    # 分别使用openpyxl和xlwt将数据写入文件

    obj_handle_excel.run_main_save_to_excel_with_openpyxl()

    obj_handle_excel.run_main_save_to_excel_with_xlwt()
    '''测试读出文件，注意openpyxl不可以读取xls的文件,xlrd不可以读取xlsx格式的文件'''

    #obj_handle_excel.read_excel_with_openpyxl("testexcel2003.xls")  # 错误写法

    #obj_handle_excel.read_excel_with_openpyxl("testexcel2003.xls") # 错误写法

    #obj_handle_excel.read_excel("testexcel2003.xls")

    #obj_handle_excel.read_excel_with_openpyxl("testexcel2007.xlsx")

# 作者：小老鼠Python零基础速学
# 链接：https://www.jianshu.com/p/6c20d64e3df4
# 来源：简书
# 著作权归作者所有。商业转载请联系作者获得授权，非商业转载请注明出处。
