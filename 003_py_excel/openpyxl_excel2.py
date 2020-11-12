# coding: utf8

from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

head_row_labels = [u'学生ID', u'学生姓名', u'联系方式', u'知识点ID', u'知识点名称', 'memo']

wb = Workbook()
ws = wb.worksheets[0]
## 写入整行
ws.append(head_row_labels)
## 写入指定单元格
ws['A2'].value = 'xx'
wb.save("test2.xlsx")
